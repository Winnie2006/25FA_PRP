import pydicom
import numpy as np
import os
import cv2
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook

class DicomViewer:
    def __init__(self, root, folder_path, output_file_path):
        self.root = root
        self.root.title("DICOM查看器")
        self.folder_path = folder_path
        self.dicom_paths = self.get_all_dicoms(folder_path)
        self.output_file_path = output_file_path

        if not self.dicom_paths:
            raise ValueError("指定的文件夹中没DICOM文件")

        # 检查 Excel 文件是否存在
        if not os.path.exists(self.output_file_path):
            self.create_excel_file()

        # 从 Excel 文件中获取当前未分类的视频索引
        self.current_index = self.find_first_unclassified_video()

        # self.video_path = self.videos[self.current_index]
        # self.cap = cv2.VideoCapture(self.video_path)
        # 改为读取所有DICOM文件并存储为图像列表
        self.images = []
        for p in self.dicom_paths:
            arr = self.load_dicom_image(p)
            if arr is not None:
                self.images.extend(arr)  # 展开成单帧列表

        if not self.images:
            raise ValueError("没有可用的视频帧 (所有文件都可能是截图/频谱)")

        self.label = tk.Label(root)
        self.label.pack()

        self.root.bind("<Left>", self.show_previous_frame)
        self.root.bind("<Right>", self.show_next_frame)
        self.root.bind("<KP_Left>", self.show_previous_frame)  # 左方向键
        self.root.bind("<KP_Right>", self.show_next_frame)  # 右方向键
        # self.root.bind("<space>", self.toggle_pause)  # 按空格键暂停或继续(现在不是自动播放，故不需要)

        # self.paused = False
        # self.update_frame()
        self.show_frame()

        # 创建分类标注窗口
        self.create_classification_window()

    # def get_all_videos(self, folder_path):
    #     videos = []
    #     for subdir, dirs, files in os.walk(folder_path):
    #         for filename in files:
    #             if filename.lower().endswith(('.mp4', '.avi', '.mov', '.mkv', '.flv')):
    #                 videos.append(os.path.join(subdir, filename))
    #     return videos

    def get_all_dicoms(self, folder_path):
        dicoms = []
        for subdir, dirs, files in os.walk(folder_path):
            for filename in files:
                # if filename.lower().endswith(".dcm"):
                #     dicoms.append(os.path.join(subdir, filename))
                filepath = os.path.join(subdir, filename)
                try:
                    ds = pydicom.dcmread(filepath, stop_before_pixels=True)
                    dicoms.append(filepath)
                except:
                    continue
        return dicoms

    def load_dicom_image(self, path):
        ds = pydicom.dcmread(path)
        pixel_array = ds.pixel_array

        # 判断是否是视频帧 (multi-frame)
        if pixel_array.ndim == 3 and pixel_array.shape[0] > 1:
            frames = []
            for frame in pixel_array:
                frame = self.normalize_image(frame)
                frames.append(frame)
            return frames
        else:
            # 单帧 -> 认为是截图/频谱，丢弃
            return None
    
    def normalize_image(self, img):
        img = img.astype(np.float32)
        img = 255 * (img - np.min(img)) / (np.max(img) - np.min(img) + 1e-5)
        return img.astype(np.uint8)

    def show_frame(self):
        if 0 <= self.current_index < len(self.images):
            frame = self.images[self.current_index]
            img = Image.fromarray(frame).convert("L")  # 灰度图
            photo = ImageTk.PhotoImage(image=img)
            self.label.config(image=photo)
            self.label.image = photo
    
    def show_previous_frame(self, event=None):
        self.current_index = (self.current_index - 1) % len(self.images)
        self.show_frame()

    def show_next_frame(self, event=None):
        self.current_index = (self.current_index + 1) % len(self.images)
        self.show_frame()

    # def update_frame(self):
    #     if not self.paused:
    #         ret, frame = self.cap.read()
    #         if ret:
    #             frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    #             img = Image.fromarray(frame)
    #             photo = ImageTk.PhotoImage(image=img)
    #             self.label.config(image=photo)
    #             self.label.image = photo
    #             self.root.after(30, self.update_frame)  # 每30ms更新一次
    #         else:
    #             self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)  # 重置视频到开头
    #             self.update_frame()

    # def show_previous_video(self, event=None):
    #     self.cap.release()
    #     self.current_index = (self.current_index - 1) % len(self.videos)
    #     self.video_path = self.videos[self.current_index]
    #     self.cap = cv2.VideoCapture(self.video_path)
    #     self.paused = False
    #     self.update_frame()

    # def show_next_video(self, event=None):
    #     self.cap.release()
    #     self.current_index = (self.current_index + 1) % len(self.videos)
    #     self.video_path = self.videos[self.current_index]
    #     self.cap = cv2.VideoCapture(self.video_path)
    #     self.paused = False
    #     self.update_frame()

    # def toggle_pause(self, event=None):
    #     self.paused = not self.paused
    #     if not self.paused:
    #         self.update_frame()

    def close(self):
        self.cap.release()
        self.root.destroy()

    def create_excel_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dicom Names"
        ws.append(["文件名", "分类"])

        for dicom_path in self.dicom_paths:
            dicom_name = os.path.basename(dicom_path)
            ws.append([dicom_name, "未分类"])

        wb.save(self.output_file_path)
        print(f"Excel 文件已创建并保存: {self.output_file_path}")

    def find_first_unclassified_video(self):
        wb = load_workbook(self.output_file_path)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=2, values_only=True), start=1):
            if row[1] == "未分类":
                return i - 1  # 返回未分类视频的索引
        return 0  # 如果没有未分类的视频，从第一个视频开始

    def find_next_unclassified_video(self):
        wb = load_workbook(self.output_file_path)
        ws = wb.active

        for i in range(self.current_index + 1, len(self.dicom_paths) + 1):
            row = ws.cell(row=i + 1, column=2).value
            if row == "未分类":
                return i - 1  # 返回未分类视频的索引
        return -1  # 所有视频都已分类

    def create_classification_window(self):
        self.classification_window = tk.Toplevel(self.root)
        self.classification_window.title("视频分类")
        self.classification_window.geometry("800x600")

        self.classification_var = tk.StringVar(self.classification_window)

        # 创建滚动条
        scrollbar = tk.Scrollbar(self.classification_window)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 创建 Canvas
        canvas = tk.Canvas(self.classification_window, yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=canvas.yview)

        # 创建一个 Frame 作为 Canvas 的子元素
        self.classification_frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=self.classification_frame, anchor='nw')

        # 大类别和小类别的嵌套关系
        categories = {
            "剑突下系列": [
                "剑突下四腔切面",
                "剑突下双房切面",
                "剑突下五腔切面",
                "剑突下主动脉瓣短轴切面",
                "剑突下右室流入-流出道切面",
                "下腔静脉长轴切面",
                "腹主动脉长轴切面",
                "腹主动脉-下腔静脉短轴切面"
            ],
            "胸骨旁长轴系列": [
                "胸骨旁长轴切面",
                "胸骨旁升主动脉长轴切面",
                "右室流入道切面",
                "右室流出道切面"
            ],
            "心尖系列": [
                "心尖四腔",
                "心尖四腔（遮挡（部分）房室腔）",
                "心尖四腔 （非标准，聚焦冠状静脉窦）",
                "心尖五腔 （标准切面，聚焦LVOT、AV）",
                "心尖五腔 （非标准，聚焦VSD）",
                "心尖两腔切面",
                "心尖三腔切面（聚焦二尖瓣）",
                "心尖三腔切面（聚焦LVOT）"
            ],
            "胸骨旁短轴系列": [
                "胸骨旁短轴（聚焦三尖瓣、RVOT、主动脉瓣、VSD、肺动脉瓣）",
                "胸骨旁短轴（聚焦右冠）",
                "胸骨旁短轴（聚焦左主干+LAD）",
                "胸骨旁短轴（聚焦LCX）",
                "胸骨旁肺动脉长轴 （聚焦肺动脉及分支）",
                "胸骨旁肺动脉长轴（聚焦PDA）",
                "胸骨旁短轴（二尖瓣水平）",
                "胸骨旁短轴（乳头肌水平）",
                "胸骨旁短轴（心尖水平）"
            ],
            "胸骨上窝系列": [
                "胸骨上窝主动脉弓短轴",
                "胸骨上窝主动脉弓短轴（非标准，聚焦肺静脉左房入口）",
                "胸骨上窝主动脉弓长轴"
            ]
        }

        for main_category, sub_categories in categories.items():
            main_frame = tk.LabelFrame(self.classification_frame, text=main_category, padx=10, pady=10)
            main_frame.pack(fill="both", expand=True)

            for sub_category in sub_categories:
                tk.Radiobutton(main_frame, text=sub_category, variable=self.classification_var, value=sub_category).pack(anchor=tk.W)

        # 更新 Canvas 和滚动区域
        self.classification_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

        self.classify_button = tk.Button(self.classification_window, text="分类", command=self.classify_video)
        self.classify_button.pack()

        self.delete_button = tk.Button(self.classification_window, text="删除分类", command=self.delete_classification)
        self.delete_button.pack()

    def classify_dicom(self):
        dicom_name = os.path.basename(self.dicom_paths[self.current_index])
        classification = self.classification_var.get()

        if not classification:
            messagebox.showwarning("警告", "请选择一个分类！")
            return

        wb = load_workbook(self.output_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_col=2, values_only=False):
            if row[0].value == dicom_name:
                row[1].value = classification
                break

        wb.save(self.output_file_path)
        print(f"视频 {dicom_name} 分类为 {classification}，已保存到 Excel 文件。")

        # 自动切换到下一个未分类的视频
        self.current_index = self.find_next_unclassified_video()
        if self.current_index == -1:
            self.save_classification_summary()
            messagebox.showinfo("完成", "所有视频分类已完成！")
        else:
            self.video_path = self.videos[self.current_index]
            self.cap = cv2.VideoCapture(self.video_path)
            self.update_frame()

    def delete_classification(self):
        video_name = os.path.basename(self.videos[self.current_index])

        wb = load_workbook(self.output_file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_col=2, values_only=False):
            if row[0].value == video_name:
                row[1].value = "未分类"
                break

        wb.save(self.output_file_path)
        print(f"视频 {video_name} 的分类已删除，当前分类为 '未分类'。")

    def save_classification_summary(self):
        wb = load_workbook(self.output_file_path)
        ws = wb.active

        classification_counts = {}
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            classification = row[1]
            if classification != "未分类":
                if classification in classification_counts:
                    classification_counts[classification] += 1
                else:
                    classification_counts[classification] = 1

        # 创建一个新的工作表来保存分类统计结果
        if "Classification Summary" in wb.sheetnames:
            summary_ws = wb["Classification Summary"]
        else:
            summary_ws = wb.create_sheet(title="Classification Summary")

        summary_ws.append(["分类", "数量"])
        for classification, count in classification_counts.items():
            summary_ws.append([classification, count])

        wb.save(self.output_file_path)
        print("分类统计结果已保存到 Excel 文件。")
      
# 获取当前用户的桌面路径
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

# 指定视频所在的根目录（假设视频文件夹也在桌面上）
dicom_folder_name = "dicoms"  # 视频文件夹的名称
dicom_folder_path = os.path.join(desktop_path, dicom_folder_name)

# 指定保存分类结果的 Excel 文件路径
excel_file_name = "dicom_names.xlsx"
output_file_path = os.path.join(desktop_path, excel_file_name)

# 创建主窗口
root = tk.Tk()
app = DicomViewer(root, dicom_folder_path, output_file_path)

# 运行主循环
root.mainloop()